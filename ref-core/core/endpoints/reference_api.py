import json

from flask.ext.api import status
import flask as fk

from core import app, CORE_URL, crossdomain, core_response
from refdb.common.models import SetModel
from refdb.common.models import ReferenceModel

import mimetypes
import json
import traceback
import datetime
import random
import string
import os
import thread
from StringIO import StringIO
from utils import Build, Evaluation
from openpyxl import load_workbook
import numpy as np

# Fit the reference mean data to extract a function for the data.

@app.route(CORE_URL + '/reference/evaluate/data/<ref_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_evaluate_data(ref_id):
    if fk.request.method == 'POST':
        _ref = ReferenceModel.objects.with_id(ref_id)
        if _ref is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this reference.')
        else:
            if fk.request.files:
                file_obj = fk.request.files['file']
                file_name = file_obj.filename
                _set, created = SetModel.objects.get_or_create(created_at=str(datetime.datetime.utcnow()))
                if created:
                    _set.filename = '{0}-{1}'.format(str(_set.id), file_name)
                    file_path = 'sets/tmp-{0}'.format(_set.filename)
                    try:
                        with open(file_path, 'wb') as set_file:
                            set_file.write(file_obj.read())
                        wb = load_workbook(file_path, read_only=True)
                        ws = wb.active
                        pressure = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                        uptake = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                        for odx, row in enumerate(ws.rows):
                            if odx >= 2:
                                # print "--- row ---"
                                if row[0].value is not None:
                                    pressure['aliq1']['run1'].append(row[0].value)
                                if row[1].value is not None:
                                    uptake['aliq1']['run1'].append(row[1].value)
                                
                                if row[3].value is not None:
                                    pressure['aliq1']['run2'].append(row[3].value)
                                if row[4].value is not None:
                                    uptake['aliq1']['run2'].append(row[4].value)
                                

                                if row[7].value is not None:
                                    pressure['aliq2']['run1'].append(row[7].value)
                                if row[8].value is not None:
                                    uptake['aliq2']['run1'].append(row[8].value)
                                
                                if row[10].value is not None:
                                    pressure['aliq2']['run2'].append(row[10].value)
                                if row[11].value is not None:
                                    uptake['aliq2']['run2'].append(row[11].value)

                        evaluation = Evaluation(eval_id=_set.filename, reference=_ref, pressure=pressure, uptake=uptake)
                        evaluation.run()
                        os.remove(file_path)
                        _set.delete()
                        return core_response(200, 'Results of new set {0} evaluated on reference[{1}].'.format(file_name, ref_id), evaluation.results)
                    except:
                        print traceback.print_exc()
                        _set.delete()
                        print "An error occured!!"
                        return core_response(204, 'Nothing created', 'An error occured.')
                else:
                    return core_response(204, 'Already exists', 'This should normaly never happened.')
            else:
                return core_response(204, 'Nothing created', 'You must a set file.')
    # else:
    #     return core_response(405, 'Method not allowed', 'This endpoint supports only a POST method.')

    return """
    <!doctype html>
    <title>Dataset Evaluation</title>
    <h1>Upload dataset</h1>
    <form action="" method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>
    """

@app.route(CORE_URL + '/reference/evaluate/plot/<ref_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_evaluate_plot(ref_id):
    if fk.request.method == 'POST':
        _ref = ReferenceModel.objects.with_id(ref_id)
        if _ref is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this reference.')
        else:
            if fk.request.files:
                file_obj = fk.request.files['file']
                file_name = file_obj.filename
                _set, created = SetModel.objects.get_or_create(created_at=str(datetime.datetime.utcnow()))
                if created:
                    _set.filename = '{0}-{1}'.format(str(_set.id), file_name)
                    file_path = '/tmp/{0}'.format(_set.filename)
                    try:
                        with open(file_path, 'wb') as set_file:
                            set_file.write(file_obj.read())
                        wb = load_workbook(file_path, read_only=True)
                        ws = wb.active
                        pressure = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                        uptake = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                        for odx, row in enumerate(ws.rows):
                            if odx >= 2:
                                # print "--- row ---"
                                if row[0].value is not None:
                                    pressure['aliq1']['run1'].append(row[0].value)
                                if row[1].value is not None:
                                    uptake['aliq1']['run1'].append(row[1].value)
                                
                                if row[3].value is not None:
                                    pressure['aliq1']['run2'].append(row[3].value)
                                if row[4].value is not None:
                                    uptake['aliq1']['run2'].append(row[4].value)
                                

                                if row[7].value is not None:
                                    pressure['aliq2']['run1'].append(row[7].value)
                                if row[8].value is not None:
                                    uptake['aliq2']['run1'].append(row[8].value)
                                
                                if row[10].value is not None:
                                    pressure['aliq2']['run2'].append(row[10].value)
                                if row[11].value is not None:
                                    uptake['aliq2']['run2'].append(row[11].value)

                        evaluation = Evaluation(eval_id=_set.filename, reference=_ref, pressure=pressure, uptake=uptake)
                        evaluation.run()
                        # print str(evals)
                        os.remove(file_path)
                        _set.delete()
                        reslt_path = evaluation.error()
                        file_buffer = None
                        try:
                            with open(reslt_path, 'r') as _file:
                                file_buffer = StringIO(_file.read())
                            file_buffer.seek(0)
                        except:
                            print traceback.print_exc()
                        if file_buffer != None:
                            # os.remove(reslt_path)
                            return fk.send_file(file_buffer, attachment_filename=reslt_path.split('/')[1], mimetype='image/png')
                        else:
                            return core_response(404, 'Request suggested an empty response', 'Unable to return plot image.')
                    except:
                        print traceback.print_exc()
                        _set.delete()
                        print "An error occured!!"
                        return core_response(204, 'Nothing created', 'An error occured.')
                else:
                    return core_response(204, 'Already exists', 'This should normaly never happened.')
            else:
                return core_response(204, 'Nothing created', 'You must a set file.')
    # else:
    #     return core_response(405, 'Method not allowed', 'This endpoint supports only a POST method.')

    return """
    <!doctype html>
    <title>Dataset Evaluation</title>
    <h1>Upload dataset</h1>
    <form action="" method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>
    """

@app.route(CORE_URL + '/reference/all', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_all():
    if fk.request.method == 'GET':
        refs = {'total_refs':0, 'refs':[]}
        for _ref in ReferenceModel.objects():
            if len(_ref.formula) == 0 and _ref.status == "done":
                z = np.polyfit(_ref.fit_pressure, _ref.mn_uptake, 30)
                _ref.formula = z.tolist()
                _ref.save()
            refs['refs'].append(_ref.summary())
        refs['total_refs'] = len(refs['refs'])
        return core_response(200, 'All the refs', refs)
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/reference/delete/<ref_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_delete(ref_id):
    if fk.request.method == 'GET':
        _ref = ReferenceModel.objects.with_id(ref_id)
        if _ref is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this reference.')
        else:
            if _ref.status != 'done':
                for _set in _ref.sets['sets']:
                    set_ = SetModel.objects.with_id(_set['id'])
                    set_.status = "unlocked"
                    set_.save()
            else:
                try:
                    ref_ = 'plots/ref-{0}.png'.format(str(_ref.id))
                    os.remove(ref_)
                    sta_ = 'plots/stats-ref-{0}.png'.format(str(_ref.id))
                    os.remove(sta_)
                    print "Reference stats deleted."
                except:
                    print traceback.print_exc()
                    print "plots deletion failed."
            _ref.delete()
            return core_response(200, 'Reference [{0}] deleted'.format(ref_id), 'The provided reference has been deleted.')
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/reference/create', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_create():
    if fk.request.method == 'GET':
        new_ref = ReferenceModel.objects(status='new').first()
        if new_ref:
            return core_response(200, 'Existing new reference found','The reference [{0}] was already created and not used.'.format(str(new_ref.id)))
        else:
            _ref, created = ReferenceModel.objects.get_or_create(created_at=str(datetime.datetime.utcnow()))
            if created:
                sets = []
                for _set in SetModel.objects():
                    if _set.status != 'excluded':
                        if str(_set.id) not in sets:
                            if _set.updated_from:
                                sets.remove(_set.updated_from)
                                sets.append(str(_set.id))
                            else:
                                sets.append(str(_set.id))
                if len(sets) > 0:
                    _sets = []
                    for _set in sets:
                        set_ = SetModel.objects.with_id(_set)
                        set_.status = "locked"
                        set_.save()
                        if set_:
                            _sets.append({'id':str(set_.id), 'filename':set_.filename})
                    _ref.sets = {'size':len(_sets), 'sets':_sets}
                    _ref.save()
                    return core_response(200, 'Creation succeed','Reference [{0}] created.'.format(str(_ref.id)))
                else:
                    return core_response(204, 'Creation failed','No data sets provided.')
            else:
                return core_response(204, 'Creation failed','Could not create a new reference.')
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/reference/build/<ref_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_build(ref_id):
    if fk.request.method == 'GET':
        _ref = ReferenceModel.objects(id=ref_id).first()
        if _ref is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this reference.')
        else:
            # Should we allow a built to be done again when already done?
            if _ref.status == "done":
                return core_response(204, 'Built failed','Reference[{0}] already built.'.format(ref_id))
            else:
                sets = []
                for _set_id in _ref.sets['sets']:
                    _set = SetModel.objects.with_id(_set_id['id'])
                    if _set:
                        sets.append(_set)
                try:
                    # build
                    # maybe run a daemon
                    build = Build(_ref)
                    build.compute()
                    build.plot_all()
                    build.plot_stat()
                    build.plot_bars()
                    for _set in sets:
                        _set.status = 'used'
                        _set.save()
                    _ref.status = 'done'
                    _ref.save()
                    return core_response(200, 'Build done.', 'Reference [{0}] was successfully built.'.format(str(_ref.id)))
                except:
                    print traceback.print_exc()
                    print "reference build failed."
                    return core_response(500, 'Error occured', 'Failed to build this reference.')
            
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/reference/sets/<ref_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_sets(ref_id):
    if fk.request.method == 'GET':
        _ref = ReferenceModel.objects(id=ref_id).first()
        if _ref is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this reference.')
        else:
            sets = []
            for _set_id in _ref.sets['sets']:
                _set = SetModel.objects.with_id(_set_id['id'])
                if _set:
                    sets.append(_set.summary())
            return core_response(200, 'Reference [{0}] sets.'.format(str(_ref.id)), sets)
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/reference/show/<ref_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def reference_show(ref_id):
    if fk.request.method == 'GET':
        _ref = ReferenceModel.objects(id=ref_id).first()
        if _ref is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this reference.')
        else:
            return core_response(200, 'Reference [{0}] info.'.format(str(_ref.id)), _ref.info())
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')
