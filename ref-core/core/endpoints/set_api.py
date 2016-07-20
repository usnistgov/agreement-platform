import json

from flask.ext.api import status
import flask as fk
from flask import flash

from core import app, CORE_URL, crossdomain, core_response
from refdb.common.models import SetModel
from refdb.common.models import ReferenceModel

import mimetypes
import json
import traceback
import datetime
import random
import string
import os
import thread
from StringIO import StringIO
from openpyxl import load_workbook

@app.route(CORE_URL + '/set/upload', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_upload():
    if fk.request.method == 'POST':
        if fk.request.files:
            file_obj = fk.request.files['file']
            file_name = file_obj.filename
            _set, created = SetModel.objects.get_or_create(created_at=str(datetime.datetime.utcnow()))
            if created:
                _set.filename = '{0}-{1}'.format(str(_set.id), file_name)
                file_path = 'sets/{0}'.format(_set.filename)
                try:
                    with open(file_path, 'wb') as set_file:
                        set_file.write(file_obj.read())
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active
                    pressure = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                    uptake = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                    for odx, row in enumerate(ws.rows):
                        if odx >= 2:
                            # print "--- row ---"
                            if row[0].value is not None:
                                pressure['aliq1']['run1'].append(row[0].value)
                            if row[1].value is not None:
                                uptake['aliq1']['run1'].append(row[1].value)
                            
                            if row[3].value is not None:
                                pressure['aliq1']['run2'].append(row[3].value)
                            if row[4].value is not None:
                                uptake['aliq1']['run2'].append(row[4].value)
                            

                            if row[7].value is not None:
                                pressure['aliq2']['run1'].append(row[7].value)
                            if row[8].value is not None:
                                uptake['aliq2']['run1'].append(row[8].value)
                            
                            if row[10].value is not None:
                                pressure['aliq2']['run2'].append(row[10].value)
                            if row[11].value is not None:
                                uptake['aliq2']['run2'].append(row[11].value)
                    _set.raw_pressure = pressure
                    _set.raw_uptake = uptake
                    print str(pressure)
                    print str(uptake)
                    _set.status = "new"
                    _set.save()
                    # return core_response(200, 'New set {0} uploaded.'.format(file_name), _set.info())
                    #Materialize.toast('New set {0} uploaded.'.format(file_name), 4000)
                    return fk.redirect('http://0.0.0.0:4000/sets/?result=success')
                except:
                    print traceback.print_exc()
                    _set.delete()
                    print "An error occured!!"
                    # return core_response(204, 'Nothing created', 'An error occured.')
                    #Materialize.toast('An error occured. Please try again later.', 4000)
                    return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
            else:
                # return core_response(204, 'Already exists', 'This should normaly never happened.')
                #Materialize.toast('Already exists! This should normaly never happened.', 4000)
                return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
        else:
            # return core_response(204, 'Nothing created', 'You must a set file.')
            #Materialize.toast('Upload failed. You must provide a set file.', 4000)
            return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
    # else:
    #     return core_response(405, 'Method not allowed', 'This endpoint supports only a POST method.')

    return """
    <!doctype html>
    <html>
        <head>
          <!-- css  -->
          <link href="https://fonts.googleapis.com/icon?family=Material+Icons" rel="stylesheet">
          <link href="http://0.0.0.0:4000/css/materialize.min.css" type="text/css" rel="stylesheet" media="screen,projection"/>
          <link href="http://0.0.0.0:4000/css/style.css" type="text/css" rel="stylesheet" media="screen,projection"/>

          <!--Let browser know website is optimized for mobile-->
          <meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1.0, user-scalable=no"/>
          <title>Reference Platform</title>
        </head>
        <body>
            <nav class="white" role="navigation">
                <div class="nav-wrapper container">
                  <a id="logo-container" href="http://0.0.0.0:4000" class="teal-text text-lighten-2">Reference</a>
                </div>
            </nav>
            <div class="valign center-align">
                <h1>Upload new dataset</h1>
                <form action="" method=post enctype=multipart/form-data>
                    <input type=file name=file>
                    <input type=submit value=Upload>
                </form>
            </div>
            <!--Import jQuery before materialize.js-->
            <script type="text/javascript" src="https://code.jquery.com/jquery-2.1.1.min.js"></script>
            <script type="text/javascript" src="http://0.0.0.0:4000/js/materialize.min.js"></script>
        </body>
    </html>
    """

@app.route(CORE_URL + '/set/change/<set_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_change(set_id):
    if fk.request.method == 'POST':
        if fk.request.files:
            file_obj = fk.request.files['file']
            file_name = file_obj.filename
            old_set = SetModel.objects.with_id(set_id)
            if old_set:
                _set, created = SetModel.objects.get_or_create(created_at=str(datetime.datetime.utcnow()))
                _set.filename = '{0}-{1}'.format(str(_set.id), file_name)
                file_path = 'sets/{0}'.format(_set.filename)
                _set.updated_from = str(set_id)
                try:
                    with open(file_path, 'wb') as set_file:
                        set_file.write(file_obj.read())
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active
                    pressure = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                    uptake = {'aliq1':{'run1':[], 'run2':[]}, 'aliq2':{'run1':[], 'run2':[]}}
                    for odx, row in enumerate(ws.rows):
                        if odx >= 2:
                            # print "--- row ---"
                            if row[0].value is not None:
                                pressure['aliq1']['run1'].append(row[0].value)
                            if row[1].value is not None:
                                uptake['aliq1']['run1'].append(row[1].value)
                            
                            if row[3].value is not None:
                                pressure['aliq1']['run2'].append(row[3].value)
                            if row[4].value is not None:
                                uptake['aliq1']['run2'].append(row[4].value)
                            

                            if row[7].value is not None:
                                pressure['aliq2']['run1'].append(row[7].value)
                            if row[8].value is not None:
                                uptake['aliq2']['run1'].append(row[8].value)
                            
                            if row[10].value is not None:
                                pressure['aliq2']['run2'].append(row[10].value)
                            if row[11].value is not None:
                                uptake['aliq2']['run2'].append(row[11].value)
                    _set.raw_pressure = pressure
                    _set.raw_uptake = uptake
                    # print str(pressure)
                    # print str(uptake)
                    _set.status = "new"
                    _set.save()
                    #Materialize.toast('New set {0} uploaded to change {1}.'.format(_set.filename, old_set.filename), 4000)
                    return fk.redirect('http://0.0.0.0:4000/sets/?result=success')
                    # return core_response(200, 'New set {0} uploaded to change {1}.'.format(_set.filename, old_set.filename), _set.info())
                except:
                    print traceback.print_exc()
                    _set.delete()
                    print "An error occured!!"
                    # return core_response(204, 'Nothing created', 'An error occured.')
                    #Materialize.toast('An error occured. Please try again later.', 4000)
                    return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
            else:
                # return core_response(204, 'Nothing changed', 'Could not find this filename')
                #Materialize.toast('Modification failed! Could not find this filename.', 4000)
                return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
        else:
            # return core_response(204, 'Nothing created', 'You must a set file.')
            #Materialize.toast('Modification failed. You must provide a set file.', 4000)
            return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
    # else:
    #     return core_response(405, 'Method not allowed', 'This endpoint supports only a POST method.')

    return """
    <!doctype html>
    <title>Data set Modification</title>
    <h1>Change dataset file</h1>
    <form action="" method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>
    """

@app.route(CORE_URL + '/set/all', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_all():
    if fk.request.method == 'GET':
        sets = {'total_sets':0, 'sets':[]}
        for _set in SetModel.objects():
            sets['sets'].append(_set.summary())
        sets['total_sets'] = len(sets['sets'])
        return core_response(200, 'All the sets', sets)
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/set/delete/<set_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_delete(set_id):
    if fk.request.method == 'GET':
        _set = SetModel.objects.with_id(set_id)
        if _set is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this set.')
        else:
            _refs = ReferenceModel.objects()
            ref = None
            for _ref in _refs:
                for set_ in _ref.sets['sets']:
                    if set_['id'] == set_id:
                        ref = _ref
                        break
            if ref:
                # flash(u'Invalid password provided', 'error')
                # Materialize.toast('Set deletion faile. The set[{0}] was used/is being used in generating a reference[{1}]. It cannot be deleted.'.format(set_id, str(ref.id)), 4000)
                return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
                # return core_response(204, 'Set deletion failed', 'The set[{0}] was used/is being used in generating a reference[{1}]. It cannot be deleted.'.format(set_id, str(ref.id)))
            else:
                file_path = 'sets/{0}'.format(_set.filename)
                try:
                    os.remove(file_path)
                    print "File deleted."
                    _set.delete()
                    # Materialize.toast('Set [{0}] deleted'.format(_set.filename), 4000)
                    return fk.redirect('http://0.0.0.0:4000/sets/?result=success')
                    # return core_response(200, 'Set [{0}] deleted'.format(_set.filename), 'The provided set has been deleted.')
                except:
                    _set.delete()
                    print traceback.print_exc()
                    print "file deletion failed."
                    # return core_response(500, 'Error occured', 'Failed to delete the set file.')
                    # Materialize.toast('An error occured. Please try again later.', 4000)
                    return fk.redirect('http://0.0.0.0:4000/sets/?result=failure')
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/set/show/<set_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_show(set_id):
    if fk.request.method == 'GET':
        _set = SetModel.objects.with_id(set_id)
        if _set is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this set.')
        else:
            return core_response(200, 'Set [{0}] info'.format(_set.filename), _set.info())
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/set/exclude/<set_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_exclude(set_id):
    if fk.request.method == 'GET':
        _set = SetModel.objects.with_id(set_id)
        if _set is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this set.')
        else:
            if _set.status != 'excluded':
                _set.status = 'excluded'
                _set.save()
                return core_response(200, 'Exclusion succeed', 'Set [{0}] is now excluded for the upcoming new reference.'.format(_set.filename))
            else:
                return core_response(204, 'Exclusion failed', 'Set [{0}] is already excluded'.format(_set.filename))
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/set/include/<set_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def set_include(set_id):
    if fk.request.method == 'GET':
        _set = SetModel.objects.with_id(set_id)
        if _set is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this set.')
        else:
            if _set.status == 'excluded':
                _set.status = 'included'
                _set.save()
                return core_response(200, 'Inclusion succeed', 'Set [{0}] is now included for the upcoming new reference.'.format(_set.filename))
            else:
                return core_response(204, 'Inclusion failed', 'Set [{0}] is not excluded'.format(_set.filename))
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

@app.route(CORE_URL + '/set/download/<set_id>', methods=['GET','POST','PUT','UPDATE','DELETE'])
@crossdomain(origin='*')
def download_file(set_id):
    if fk.request.method == 'GET':
        _set = SetModel.objects.with_id(set_id)
        if _set is None:
            return core_response(404, 'Request suggested an empty response', 'Unable to find this set.')
        else:
            file_buffer = None
            try:
                file_path = 'sets/{0}'.format(_set.filename)
                with open(file_path, 'r') as _file:
                    file_buffer = StringIO(_file.read())
                file_buffer.seek(0)
            except:
                print traceback.print_exc()
                pass
            if file_buffer != None:
                return fk.send_file(file_buffer, as_attachment=True, attachment_filename=_set.filename, mimetype='application/zip')
            else:
                return core_response(404, 'Request suggested an empty response', 'Unable to find this set file.')
    else:
        return core_response(405, 'Method not allowed', 'This endpoint supports only a GET method.')

